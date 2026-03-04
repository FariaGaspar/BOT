"""
Script para enviar planeamento para o ficheiro Wialong V5.2
Este script atualiza o separador "Planeamento" do ficheiro Wialong com as encomendas atribuídas
"""
import sqlite3
import sys
from datetime import date
import json
from io import BytesIO

# Tentar importar openpyxl (biblioteca para trabalhar com Excel)
OPENPYXL_AVAILABLE = False
OPENPYXL_VBA_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
    # Verificar se suporta VBA (macros)
    try:
        # Testar se keep_vba está disponível
        OPENPYXL_VBA_AVAILABLE = True
    except:
        OPENPYXL_VBA_AVAILABLE = False
    print("✓ openpyxl disponível")
except ImportError:
    OPENPYXL_AVAILABLE = False
    OPENPYXL_VBA_AVAILABLE = False
    print("⚠ AVISO: openpyxl não está instalado.")
    print("   Para instalar, execute: pip install openpyxl")
    print("   Ou: pip install -r requirements.txt")

# Caminho do banco de dados (relativo ao diretório do script)
import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(SCRIPT_DIR, 'planeamento.db')

def encontrar_ficheiro_wialong():
    """Tentar encontrar o ficheiro Wialong em locais comuns"""
    import os
    
    # Pasta Wialon junto à aplicação (prioridade máxima)
    pasta_aplicacao = os.path.dirname(os.path.abspath(__file__))
    pasta_wialon_aplicacao = os.path.join(pasta_aplicacao, "Wialon")
    
    # Desktop do utilizador
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    pasta_wialon_desktop = os.path.join(desktop, "Wialon")
    
    locais_comuns = [
        # Pasta Wialon junto à aplicação (prioridade máxima)
        os.path.join(pasta_wialon_aplicacao, "Wialong.xlsx"),
        os.path.join(pasta_wialon_aplicacao, "Wialong.xlsm"),
        os.path.join(pasta_wialon_aplicacao, "Wialong V5.2.xlsx"),
        os.path.join(pasta_wialon_aplicacao, "Wialong V5.2.xlsm"),
        # Pasta Wialon no Desktop
        os.path.join(pasta_wialon_desktop, "Wialong.xlsx"),
        os.path.join(pasta_wialon_desktop, "Wialong.xlsm"),
        os.path.join(pasta_wialon_desktop, "Wialong V5.2.xlsx"),
        os.path.join(pasta_wialon_desktop, "Wialong V5.2.xlsm"),
        # Diretamente na pasta da aplicação
        os.path.join(pasta_aplicacao, "Wialong.xlsx"),
        os.path.join(pasta_aplicacao, "Wialong.xlsm"),
        os.path.join(pasta_aplicacao, "Wialong V5.2.xlsx"),
        os.path.join(pasta_aplicacao, "Wialong V5.2.xlsm"),
        # Desktop
        os.path.join(desktop, "Wialong.xlsx"),
        os.path.join(desktop, "Wialong.xlsm"),
        os.path.join(desktop, "Wialong V5.2.xlsx"),
        os.path.join(desktop, "Wialong V5.2.xlsm"),
        # Documents
        os.path.join(os.path.expanduser("~"), "Documents", "Wialong.xlsx"),
        os.path.join(os.path.expanduser("~"), "Documents", "Wialong.xlsm"),
        os.path.join(os.path.expanduser("~"), "Documents", "Wialong V5.2.xlsx"),
        os.path.join(os.path.expanduser("~"), "Documents", "Wialong V5.2.xlsm"),
    ]
    
    print(f"Procurando ficheiro Wialong em {len(locais_comuns)} locais...")
    for local in locais_comuns:
        if os.path.exists(local):
            print(f"✓ Ficheiro encontrado: {local}")
            return local
        else:
            print(f"  - Não encontrado: {local}")
    
    print("✗ Ficheiro Wialong não encontrado em nenhum local")
    return None

def obter_dados_planeamento(data_planeamento):
    """Obter dados do planeamento para uma data específica, respeitando a ordem do planeamento."""
    # Usar caminho absoluto do banco de dados
    db_path = DATABASE
    if not os.path.exists(db_path):
        # Tentar caminho alternativo (pasta web_app)
        alt_path = os.path.join(SCRIPT_DIR, 'planeamento.db')
        if os.path.exists(alt_path):
            db_path = alt_path
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            vm.matricula,
            vm.codigo,
            vm.nome_motorista,
            COALESCE(pp.local_descarga, pp.cliente, '') as local_descarga,
            pp.local_carga,
            pp.material,
            ev.ordem,
            ev.id as encomenda_viatura_id
        FROM encomenda_viatura ev
        INNER JOIN viatura_motorista vm ON ev.viatura_motorista_id = vm.id
        INNER JOIN pedidos_pendentes pp ON ev.pedido_tipo = 'P' AND ev.pedido_id = pp.id
        WHERE ev.data_associacao = ?
        -- Ordenar primeiro pela ordem do motorista no painel (se existir),
        -- depois pela matrícula (estável), e finalmente pela ordem de serviço
        -- Se ordem for NULL, usar o ID como fallback para manter ordem de inserção
        ORDER BY 
            COALESCE(vm.ordem, 9999) ASC,
            vm.matricula ASC,
            COALESCE(ev.ordem, ev.id) ASC
    ''', (data_planeamento,))
    
    dados = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return dados

def atualizar_wialong(caminho_wialong, data_planeamento, limpar_antigos=True, modo_silencioso=False):
    """Atualizar ficheiro Wialong com dados do planeamento"""
    if not OPENPYXL_AVAILABLE:
        error_msg = "ERRO: openpyxl não está disponível. Não é possível atualizar o ficheiro Excel.\n\nExecute no terminal: pip install openpyxl\n\nOu instale todas as dependências: pip install -r requirements.txt"
        print(error_msg)
        if modo_silencioso:
            raise Exception(error_msg)
        return False
    
    try:
        # Verificar se o ficheiro existe
        import os
        if not os.path.exists(caminho_wialong):
            print(f"ERRO: Ficheiro não encontrado: {caminho_wialong}")
            print("\nDica: Verifique o caminho e certifique-se de que o ficheiro existe.")
            return False
        
        # Carregar ficheiro Wialong preservando TUDO (formatações, estilos, macros)
        print(f"Abrindo ficheiro: {caminho_wialong}")
        print(f"⚠ IMPORTANTE: Usando APENAS este ficheiro. Se houver outros ficheiros Wialong na pasta, serão ignorados.")
        try:
            is_xlsm = caminho_wialong.lower().endswith('.xlsm')
            if is_xlsm:
                try:
                    # Carregar preservando macros e formatações
                    wb = load_workbook(caminho_wialong, keep_vba=True, data_only=False)
                    print("✓ Ficheiro .xlsm carregado (macros e formatações preservadas)")
                except Exception as e_vba:
                    print(f"AVISO: Não foi possível carregar com macros: {e_vba}")
                    wb = load_workbook(caminho_wialong, data_only=False)
                    print("✓ Ficheiro carregado (formatações preservadas)")
            else:
                wb = load_workbook(caminho_wialong, data_only=False)
                print("✓ Ficheiro carregado (formatações preservadas)")
        except PermissionError as e:
            error_msg = f"ERRO de permissão: {e}\nO ficheiro pode estar aberto no Excel ou não tem permissões de escrita."
            print(error_msg)
            if modo_silencioso:
                raise Exception(error_msg)
            return False
        except Exception as e:
            import traceback
            error_msg = f"ERRO ao abrir ficheiro: {e}"
            print(error_msg)
            print(traceback.format_exc())
            if modo_silencioso:
                raise Exception(f"{error_msg}\nCertifique-se de que o ficheiro não está aberto no Excel.")
            return False
        
        # Verificar se existe o separador "Planeamento"
        if 'Planeamento' not in wb.sheetnames:
            if modo_silencioso:
                # Modo silencioso: criar automaticamente
                ws = wb.create_sheet('Planeamento')
            else:
                # Modo interativo: perguntar
                print(f"\nAVISO: Separador 'Planeamento' não encontrado.")
                print(f"Separadores disponíveis: {', '.join(wb.sheetnames)}")
                resposta = input("Deseja criar o separador 'Planeamento'? (s/n): ").strip().lower()
                if resposta == 's':
                    ws = wb.create_sheet('Planeamento')
                    print("✓ Separador 'Planeamento' criado.")
                else:
                    print("Operação cancelada.")
                    return False
        else:
            ws = wb['Planeamento']
            print(f"✓ Separador 'Planeamento' encontrado e selecionado.")
            print(f"  Nome do separador ativo: {ws.title}")
            print(f"  Linhas no separador: {ws.max_row}")

        # Garantir que o ficheiro final fica desbloqueado (sem proteção de folha / livro)
        try:
            # Desbloquear folha "Planeamento"
            ws.protection.sheet = False
            ws.protection.enable()
            ws.protection.disable()
        except Exception:
            # Se por algum motivo não conseguir, continuar mesmo assim
            pass
        try:
            # Desbloquear estrutura do livro (caso esteja protegida)
            if hasattr(wb, "security") and wb.security is not None:
                wb.security.lockStructure = False
                wb.security.lockWindows = False
        except Exception:
            pass
        
        # Obter dados
        if not modo_silencioso:
            print(f"\nBuscando dados para a data: {data_planeamento}")
        dados = obter_dados_planeamento(data_planeamento)
        
        if not dados:
            error_msg = f"Nenhuma encomenda atribuída para a data {data_planeamento}"
            print(f"⚠ {error_msg}")
            if modo_silencioso:
                raise Exception(error_msg)
            return False
        
        # NÃO limpar dados antigos - vamos adicionar abaixo da linha da matrícula/motorista
        print(f"Separador 'Planeamento' tem {ws.max_row} linhas. Vou adicionar logo abaixo da linha da matrícula de cada motorista.")
        
        # Criar um dicionário para mapear motoristas existentes e a linha onde está a matrícula
        motoristas_existentes = {}  # {matricula: linha_da_matricula}
        
        # Procurar motoristas existentes no ficheiro
        # Procurar a matrícula nas colunas A, B ou C (colunas 1, 2, 3)
        for linha in range(1, ws.max_row + 1):
            # Verificar coluna A (1), B (2) e C (3)
            for col in [1, 2, 3]:
                valor_celula = ws.cell(row=linha, column=col).value
                if valor_celula and isinstance(valor_celula, str):
                    valor = valor_celula.strip()
                    # Verificar se este valor corresponde a uma matrícula (comparar com as matrículas dos dados)
                    # Mas primeiro, vamos procurar pela matrícula exata que temos nos dados
                    # Vamos fazer isso depois de agrupar os dados
                    # Por agora, guardar qualquer valor não vazio na coluna 1 como possível matrícula
                    if col == 1 and valor:
                        # Guardar apenas a primeira ocorrência (linha da matrícula)
                        if valor not in motoristas_existentes:
                            motoristas_existentes[valor] = linha
        
        print(f"Valores encontrados na coluna A: {list(motoristas_existentes.keys())[:10]}...")  # Mostrar apenas os primeiros 10
        
        # Agrupar por motorista (usar nome do motorista como chave)
        # IMPORTANTE: Preservar a ordem EXATA que vem da query SQL (já ordenada por ev.ordem)
        # Não reordenar - usar a ordem que já vem correta da base de dados
        dados_agrupados = {}
        for item in dados:
            nome_motorista = item['nome_motorista'].strip() if item['nome_motorista'] else ''
            if nome_motorista and nome_motorista not in dados_agrupados:
                dados_agrupados[nome_motorista] = {
                    'nome_motorista': nome_motorista,
                    'encomendas': []
                }
            if nome_motorista:
                # Adicionar na ordem que vem da query (já ordenada por ev.ordem)
                # Sequência no ficheiro: Local de descarga - Local de carga - Material
                dados_agrupados[nome_motorista]['encomendas'].append({
                    'local_descarga': item.get('local_descarga') or '',
                    'local_carga': item.get('local_carga') or '',
                    'material': item.get('material') or '',
                    'ordem': item.get('ordem', 9999)  # Guardar para debug, mas não reordenar
                })
        
        # NÃO reordenar - a ordem já está correta da query SQL
        # A query já ordena por: vm.ordem, vm.matricula, ev.ordem
        
        # Procurar pelo nome do motorista no ficheiro
        motoristas_mapeados = {}  # {nome_motorista: linha_da_matricula}
        
        # Lista de todos os nomes de motoristas que temos nos dados
        nomes_motoristas = [v['nome_motorista'].strip() for v in dados_agrupados.values()]
        
        # Procurar cada nome de motorista nas colunas A, B ou C
        print(f"Procurando nomes de motoristas no separador 'Planeamento' (linhas 1 a {ws.max_row})...")
        print(f"Nomes a procurar: {nomes_motoristas}")
        
        for linha in range(1, ws.max_row + 1):
            for col in [1, 2, 3]:
                valor_celula = ws.cell(row=linha, column=col).value
                if valor_celula:
                    # Converter para string e limpar (remover espaços extras)
                    if isinstance(valor_celula, str):
                        valor = valor_celula.strip().upper()  # Converter para maiúsculas para comparação
                    else:
                        valor = str(valor_celula).strip().upper()
                    
                    # Verificar se este valor corresponde a um dos nomes de motoristas que temos
                    for nome_motorista in nomes_motoristas:
                        nome_upper = nome_motorista.upper().strip()
                        # Comparação exata ou se o nome está contido no valor (ou vice-versa)
                        if valor == nome_upper or nome_upper in valor or valor in nome_upper:
                            if nome_motorista not in motoristas_mapeados:
                                motoristas_mapeados[nome_motorista] = linha
                                print(f"  ✓ Motorista '{nome_motorista}' encontrado na linha {linha}, coluna {col} (valor no Excel: '{valor_celula}')")
                                break
        
        # Escrever dados abaixo de cada motorista
        # IMPORTANTE: Iterar na ordem que os motoristas aparecem nos dados originais (já ordenados)
        # Criar lista ordenada de motoristas baseada na ordem de aparição nos dados
        ordem_motoristas = []
        motoristas_vistos = set()
        for item in dados:
            nome_motorista = item['nome_motorista'].strip() if item['nome_motorista'] else ''
            if nome_motorista and nome_motorista not in motoristas_vistos:
                ordem_motoristas.append(nome_motorista)
                motoristas_vistos.add(nome_motorista)
        
        print(f"\nEscrevendo {len(dados)} encomendas de {len(dados_agrupados)} motoristas...")
        print(f"Ordem dos motoristas: {ordem_motoristas}")
        
        linhas_adicionadas = 0
        
        # Iterar na ordem correta dos motoristas (não usar .values() que pode não manter ordem)
        for nome_motorista in ordem_motoristas:
            if nome_motorista not in dados_agrupados:
                continue
            motorista_data = dados_agrupados[nome_motorista]
            nome_motorista = motorista_data['nome_motorista'].strip()
            
            # Debug: mostrar ordem das encomendas antes de escrever
            print(f"\n  📋 Motorista '{nome_motorista}' - Ordem das encomendas que serão escritas:")
            for idx, enc in enumerate(motorista_data['encomendas'], 1):
                ordem_val = enc.get('ordem', 'N/A')
                local_desc = enc.get('local_descarga', '')
                local_carga = enc.get('local_carga', '')
                print(f"    {idx}. Local descarga: '{local_desc}' | Local carga: '{local_carga}' | Ordem BD: {ordem_val}")
            
            # Verificar se o motorista já existe no ficheiro
            if nome_motorista in motoristas_mapeados:
                # Encontrar a linha onde está o nome do motorista
                linha_motorista = motoristas_mapeados[nome_motorista]
                # Adicionar logo na linha seguinte (logo abaixo do nome)
                linha_atual = linha_motorista + 1
                print(f"  ✓ Motorista '{nome_motorista}' encontrado na linha {linha_motorista}")
                print(f"    → Colando {len(motorista_data['encomendas'])} encomendas a partir da linha {linha_atual}")
            else:
                # Motorista não existe, adicionar no final
                linha_atual = ws.max_row + 1
                print(f"  ⚠ Motorista '{nome_motorista}' NÃO encontrado no ficheiro!")
                print(f"    → Adicionando no final (linha {linha_atual})")
                print(f"    → Dica: Verifique se o nome '{nome_motorista}' existe nas colunas A, B ou C do separador 'Planeamento'")
            
            # Escrever encomendas nas linhas seguintes (sem inserir linhas, apenas escrever)
            num_encomendas = len(motorista_data['encomendas'])
            if num_encomendas > 0:
                linha_escrever = linha_atual
                
                # Escrever APENAS o valor como texto, SEM tocar em formatações
                # IMPORTANTE: Escrever na ordem exata que vem da lista (já ordenada pela query)
                for encomenda in motorista_data['encomendas']:
                    # Obter células existentes (ou criar se não existirem)
                    # Sequência: Col A = Local de descarga, Col B = Local de carga, Col C = Material
                    cell_a = ws.cell(row=linha_escrever, column=1)
                    cell_b = ws.cell(row=linha_escrever, column=2)
                    cell_c = ws.cell(row=linha_escrever, column=3)
                    
                    # Apenas alterar o valor - preservar TUDO o resto (font, fill, border, alignment, etc.)
                    valor_local_descarga = str(encomenda.get('local_descarga', '') or '') or ''
                    valor_local_carga = str(encomenda.get('local_carga', '') or '') or ''
                    valor_material = str(encomenda.get('material', '') or '') or ''
                    
                    # Guardar propriedades existentes antes de alterar valor
                    # (openpyxl preserva automaticamente, mas vamos garantir)
                    cell_a.value = valor_local_descarga
                    cell_b.value = valor_local_carga
                    cell_c.value = valor_material
                    
                    # NÃO alterar: cell_a.font, cell_a.fill, cell_a.border, cell_a.alignment, etc.
                    # Apenas o valor foi alterado
                    
                    linha_escrever += 1
                    linhas_adicionadas += 1
        
        print(f"✓ Dados escritos: {linhas_adicionadas} linhas adicionadas")
        
        # Guardar o ficheiro
        print("\nGuardando ficheiro...")
        try:
            import time
            max_tentativas = 3
            tentativa = 0
            guardado = False
            
            while tentativa < max_tentativas and not guardado:
                try:
                    # Salvar o ficheiro (openpyxl preserva formatações automaticamente)
                    wb.save(caminho_wialong)
                    print(f"✓ Ficheiro guardado: {caminho_wialong}")
                    guardado = True
                except PermissionError as e_perm:
                    tentativa += 1
                    if tentativa < max_tentativas:
                        print(f"Tentativa {tentativa}/{max_tentativas}: Ficheiro pode estar em uso. Aguardando 1 segundo...")
                        time.sleep(1)
                    else:
                        error_msg = f"ERRO de permissão após {max_tentativas} tentativas: {e_perm}\nO ficheiro pode estar aberto no Excel."
                        print(error_msg)
                        if modo_silencioso:
                            raise Exception(error_msg)
                        return False
                except Exception as e_save:
                    error_msg = f"ERRO ao guardar: {e_save}"
                    print(error_msg)
                    import traceback
                    print(traceback.format_exc())
                    if modo_silencioso:
                        raise Exception(error_msg)
                    return False
                    
        except Exception as e:
            import traceback
            error_msg = f"ERRO ao guardar ficheiro: {e}"
            print(error_msg)
            print(traceback.format_exc())
            if modo_silencioso:
                raise Exception(error_msg)
            return False
        
        # Abrir ficheiro automaticamente (apenas uma vez)
        if modo_silencioso:
            try:
                import time
                import subprocess
                import platform
                time.sleep(0.5)
                print(f"Abrindo ficheiro: {caminho_wialong}")
                if platform.system() == 'Windows':
                    import os
                    # Tentar abrir com separador específico usando VBScript
                    try:
                        import tempfile
                        vbs_script = f'''
Set objExcel = CreateObject("Excel.Application")
objExcel.Visible = True
Set objWorkbook = objExcel.Workbooks.Open("{caminho_wialong.replace(chr(92), chr(92)+chr(92))}")
On Error Resume Next
objWorkbook.Worksheets("Planeamento").Activate
If Err.Number <> 0 Then
    objWorkbook.Worksheets(1).Activate
End If
On Error Goto 0
'''
                        with tempfile.NamedTemporaryFile(mode='w', suffix='.vbs', delete=False, encoding='utf-8') as f:
                            f.write(vbs_script)
                            vbs_path = f.name
                        
                        os.system(f'cscript //nologo "{vbs_path}"')
                        time.sleep(0.5)
                        try:
                            os.unlink(vbs_path)
                        except:
                            pass
                        print(f"✓ Ficheiro aberto com separador 'Planeamento'")
                    except Exception as e_vbs:
                        os.startfile(caminho_wialong)
                        print(f"✓ Ficheiro aberto")
                else:
                    subprocess.Popen([caminho_wialong], shell=True)
                    print(f"✓ Comando para abrir ficheiro executado")
            except Exception as e:
                import traceback
                print(f"ERRO ao abrir ficheiro: {e}")
                print(traceback.format_exc())
        
        if not modo_silencioso:
            print("\n" + "=" * 60)
            print("✓ FICHEIRO WIALONG ATUALIZADO COM SUCESSO!")
            print("=" * 60)
            print(f"  Data: {data_planeamento}")
            print(f"  Total de encomendas: {len(dados)}")
            print(f"  Total de viaturas: {len(dados_agrupados)}")
            print(f"  Ficheiro: {caminho_wialong}")
            print("=" * 60)
        else:
            print(f"✓ Atualização concluída: {len(dados)} encomendas, {len(dados_agrupados)} viaturas")
        return True
        
    except FileNotFoundError:
        print(f"\nERRO: Ficheiro não encontrado: {caminho_wialong}")
        print("\nVerifique:")
        print("  - O caminho está correto?")
        print("  - O ficheiro existe?")
        print("  - O ficheiro não está aberto no Excel?")
        return False
    except PermissionError:
        print(f"\nERRO: Não tem permissão para escrever no ficheiro.")
        print("\nVerifique:")
        print("  - O ficheiro não está aberto no Excel?")
        print("  - Tem permissões de escrita na pasta?")
        return False
    except Exception as e:
        print(f"\nERRO ao atualizar ficheiro: {e}")
        import traceback
        traceback.print_exc()
        return False

def atualizar_wialong_memoria(ficheiro_bytes, nome_ficheiro, data_planeamento):
    """
    Atualizar ficheiro Wialong a partir de bytes em memória
    Retorna o ficheiro atualizado em BytesIO ou None em caso de erro
    """
    if not OPENPYXL_AVAILABLE:
        print("ERRO: openpyxl não está disponível.")
        return None
    
    try:
        # Carregar ficheiro a partir de bytes
        ficheiro_io = BytesIO(ficheiro_bytes)
        
        # Verificar se é .xlsm (com macros)
        is_xlsm = nome_ficheiro.lower().endswith('.xlsm')
        
        try:
            if is_xlsm:
                try:
                    wb = load_workbook(ficheiro_io, keep_vba=True, data_only=False)
                except:
                    ficheiro_io.seek(0)  # Resetar para o início
                    wb = load_workbook(ficheiro_io, data_only=False)
            else:
                wb = load_workbook(ficheiro_io, data_only=False)
        except Exception as e:
            print(f"ERRO ao carregar ficheiro: {e}")
            return None
        
        # Verificar se existe o separador "Planeamento"
        if 'Planeamento' not in wb.sheetnames:
            ws = wb.create_sheet('Planeamento')
        else:
            ws = wb['Planeamento']

        # Garantir que a folha / livro enviados ao utilizador vêm desbloqueados
        try:
            ws.protection.sheet = False
            ws.protection.enable()
            ws.protection.disable()
        except Exception:
            pass
        try:
            if hasattr(wb, "security") and wb.security is not None:
                wb.security.lockStructure = False
                wb.security.lockWindows = False
        except Exception:
            pass
        
        # Obter dados
        dados = obter_dados_planeamento(data_planeamento)
        if not dados:
            return None
        
        # Agrupar por viatura (mesma lógica da função original)
        dados_agrupados = {}
        for item in dados:
            chave = f"{item['matricula']} - {item['nome_motorista']}"
            if chave not in dados_agrupados:
                dados_agrupados[chave] = {
                    'matricula': item['matricula'],
                    'codigo': item['codigo'],
                    'nome_motorista': item['nome_motorista'],
                    'encomendas': []
                }
            # Sequência no ficheiro: Local de descarga - Local de carga - Material
            dados_agrupados[chave]['encomendas'].append({
                'local_descarga': item.get('local_descarga') or '',
                'local_carga': item.get('local_carga') or '',
                'material': item.get('material') or '',
                'ordem': item.get('ordem', 9999)
            })
        
        # Procurar motoristas no ficheiro
        motoristas_mapeados = {}
        matriculas_dados = [v['matricula'].strip() for v in dados_agrupados.values()]
        
        for linha in range(1, ws.max_row + 1):
            for col in [1, 2, 3]:
                valor_celula = ws.cell(row=linha, column=col).value
                if valor_celula:
                    if isinstance(valor_celula, str):
                        valor = valor_celula.strip().upper()
                    else:
                        valor = str(valor_celula).strip().upper()
                    
                    for matricula in matriculas_dados:
                        matricula_upper = matricula.upper().strip()
                        if valor == matricula_upper or matricula_upper in valor or valor in matricula_upper:
                            if matricula not in motoristas_mapeados:
                                motoristas_mapeados[matricula] = linha
                            break
        
        # Iterar na ordem correta dos motoristas
        ordem_motoristas = []
        motoristas_vistos = set()
        for item in dados:
            chave = f"{item['matricula']} - {item['nome_motorista']}"
            if chave not in motoristas_vistos:
                ordem_motoristas.append(chave)
                motoristas_vistos.add(chave)
        
        # Escrever dados - APENAS valores de texto, SEM alterar NADA mais
        # Não inserir linhas, não copiar formatações, apenas escrever texto
        linhas_adicionadas = 0
        for chave_viatura in ordem_motoristas:
            if chave_viatura not in dados_agrupados:
                continue
            viatura = dados_agrupados[chave_viatura]
            matricula = viatura['matricula'].strip()
            
            if matricula in motoristas_mapeados:
                linha_atual = motoristas_mapeados[matricula] + 1
            else:
                linha_atual = ws.max_row + 1
            
            num_encomendas = len(viatura['encomendas'])
            if num_encomendas > 0:
                # Escrever diretamente nas células - NÃO inserir linhas, NÃO copiar formatações
                linha_escrever = linha_atual
                for encomenda in viatura['encomendas']:
                    # Sequência: Col 1 = Local de descarga, Col 2 = Local de carga, Col 3 = Material
                    ws.cell(row=linha_escrever, column=1, value=str(encomenda.get('local_descarga') or '') or '')
                    ws.cell(row=linha_escrever, column=2, value=str(encomenda.get('local_carga') or '') or '')
                    ws.cell(row=linha_escrever, column=3, value=str(encomenda.get('material') or '') or '')
                    
                    linha_escrever += 1
                    linhas_adicionadas += 1
        
        # Guardar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        import traceback
        print(f"ERRO ao processar ficheiro em memória: {e}")
        print(traceback.format_exc())
        return None

def main():
    """Função principal"""
    if len(sys.argv) < 3:
        print("Uso: python enviar_para_wialong.py <caminho_ficheiro_wialong> <data>")
        print("Exemplo: python enviar_para_wialong.py \"C:\\Users\\joao.gaspar\\Desktop\\Wialong V5.2.xlsx\" 2025-01-15")
        print("\nOu use sem argumentos para modo interativo:")
        sys.exit(1)
    
    caminho_wialong = sys.argv[1]
    data_planeamento = sys.argv[2]
    
    sucesso = atualizar_wialong(caminho_wialong, data_planeamento)
    sys.exit(0 if sucesso else 1)

if __name__ == '__main__':
    if len(sys.argv) == 1:
        # Modo interativo
        print("=" * 60)
        print("ENVIAR PLANEAMENTO PARA WIALONG V5.2")
        print("=" * 60)
        print()
        
        # Tentar encontrar o ficheiro Wialong em locais comuns
        import os
        locais_comuns = [
            os.path.join(os.path.expanduser("~"), "Desktop", "Wialong V5.2.xlsx"),
            os.path.join(os.path.expanduser("~"), "Desktop", "Wialong V5.2.xlsm"),
            os.path.join(os.path.expanduser("~"), "Documents", "Wialong V5.2.xlsx"),
            os.path.join(os.path.expanduser("~"), "Documents", "Wialong V5.2.xlsm"),
        ]
        
        caminho_sugerido = None
        for local in locais_comuns:
            if os.path.exists(local):
                caminho_sugerido = local
                break
        
        if caminho_sugerido:
            print(f"Ficheiro encontrado: {caminho_sugerido}")
            resposta = input("Usar este ficheiro? (s/n) [s]: ").strip().lower()
            if resposta != 'n':
                caminho = caminho_sugerido
            else:
                caminho = input("Caminho do ficheiro Wialong V5.2: ").strip().strip('"')
        else:
            print("Ficheiro Wialong não encontrado automaticamente.")
            caminho = input("Caminho do ficheiro Wialong V5.2: ").strip().strip('"')
        
        data = input("Data do planeamento (YYYY-MM-DD) [hoje]: ").strip()
        
        if not data:
            data = date.today().isoformat()
        
        print()
        sucesso = atualizar_wialong(caminho, data)
        
        if sucesso:
            print("\nPressione Enter para sair...")
            input()
    else:
        main()

